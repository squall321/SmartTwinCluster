"""
Reports Export Module
PDF 및 Excel 리포트 생성
"""

from io import BytesIO
from datetime import datetime
import json
import os

# PDF 생성
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.platypus import Image as RLImage
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Excel 생성
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

class ReportExporter:
    """리포트 내보내기 클래스"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_korean_font()
        self._setup_custom_styles()
    
    def _setup_korean_font(self):
        """한글 폰트 설정"""
        try:
            # 사용 가능한 한글 폰트 경로 목록
            font_paths = [
                '/usr/share/fonts/truetype/nanum/NanumGothic.ttf',
                '/usr/share/fonts/truetype/nanum/NanumBarunGothic.ttf',
                '/System/Library/Fonts/AppleGothic.ttf',  # macOS
                'C:/Windows/Fonts/malgun.ttf',  # Windows
                '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',  # Fallback
            ]
            
            font_registered = False
            for font_path in font_paths:
                if os.path.exists(font_path):
                    try:
                        pdfmetrics.registerFont(TTFont('Korean', font_path))
                        self.korean_font = 'Korean'
                        font_registered = True
                        print(f"✅ Korean font registered: {font_path}")
                        break
                    except Exception as e:
                        print(f"⚠️  Failed to register font {font_path}: {e}")
                        continue
            
            if not font_registered:
                print("⚠️  No Korean font found, using default (text may not display correctly)")
                self.korean_font = 'Helvetica'
        except Exception as e:
            print(f"⚠️  Font setup error: {e}")
            self.korean_font = 'Helvetica'
    
    def _setup_custom_styles(self):
        """커스텀 스타일 설정"""
        # 제목 스타일
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontName=self.korean_font,
            fontSize=24,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # 부제목 스타일
        self.subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=self.styles['Heading2'],
            fontName=self.korean_font,
            fontSize=16,
            textColor=colors.HexColor('#3b82f6'),
            spaceAfter=12,
            spaceBefore=20
        )
        
        # 본문 스타일
        self.body_style = ParagraphStyle(
            'CustomBody',
            parent=self.styles['Normal'],
            fontName=self.korean_font,
            fontSize=10,
            spaceAfter=12
        )
    
    # ============================================
    # PDF 생성
    # ============================================
    
    def generate_usage_report_pdf(self, data):
        """사용량 리포트 PDF 생성"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        
        # 제목
        title = Paragraph("리소스 사용량 리포트", self.title_style)
        story.append(title)
        story.append(Spacer(1, 0.2 * inch))
        
        # 기간 정보
        period_text = f"기간: {data['period']['start']} ~ {data['period']['end']} ({data['period']['days']}일)"
        period = Paragraph(period_text, self.body_style)
        story.append(period)
        story.append(Spacer(1, 0.3 * inch))
        
        # 요약 섹션
        summary_title = Paragraph("📊 사용량 요약", self.subtitle_style)
        story.append(summary_title)
        
        # 요약 테이블
        summary_data = [
            ['항목', '사용량'],
            ['CPU 시간', f"{data['total']['cpu_hours']:.2f} hours"],
            ['GPU 시간', f"{data['total']['gpu_hours']:.2f} hours"],
            ['메모리', f"{data['total']['memory_gb_hours']:.2f} GB·hours"],
            ['제출된 작업', f"{data['total']['jobs_submitted']}개"],
            ['완료된 작업', f"{data['total']['jobs_completed']}개"],
            ['실패한 작업', f"{data['total']['jobs_failed']}개"],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), self.korean_font),  # 한글 폰트 적용
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3 * inch))
        
        # 비용 섹션
        cost_title = Paragraph("💰 비용 분석", self.subtitle_style)
        story.append(cost_title)
        
        cost_data = [
            ['항목', '비용 (USD)'],
            ['CPU 비용', f"${data['costs']['cpu_cost']:.2f}"],
            ['GPU 비용', f"${data['costs']['gpu_cost']:.2f}"],
            ['메모리 비용', f"${data['costs']['memory_cost']:.2f}"],
            ['총 비용', f"${data['costs']['total_cost']:.2f}"],
        ]
        
        cost_table = Table(cost_data, colWidths=[3*inch, 3*inch])
        cost_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), self.korean_font),  # 한글 폰트 적용
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fef3c7')),
            ('FONTNAME', (0, -1), (-1, -1), self.korean_font),  # 총 비용 굵게
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(cost_table)
        story.append(Spacer(1, 0.3 * inch))
        
        # 일별 데이터 (최근 7일)
        daily_title = Paragraph("📅 일별 상세 데이터", self.subtitle_style)
        story.append(daily_title)
        
        daily_data = [['날짜', 'CPU', 'GPU', '작업 완료']]
        for day in data['daily_data'][-7:]:  # 최근 7일만
            daily_data.append([
                day['date'],
                f"{day['cpu_hours']:.1f}h",
                f"{day['gpu_hours']:.1f}h",
                f"{day['jobs_completed']}개"
            ])
        
        daily_table = Table(daily_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        daily_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), self.korean_font),  # 한글 폰트 적용
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        story.append(daily_table)
        
        # 푸터
        story.append(Spacer(1, 0.5 * inch))
        footer_text = f"생성 시간: {data['generated_at']}"
        footer = Paragraph(footer_text, ParagraphStyle('Footer', parent=self.body_style, fontName=self.korean_font, fontSize=8, textColor=colors.grey, alignment=TA_RIGHT))
        story.append(footer)
        
        # PDF 생성
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def generate_costs_report_pdf(self, data):
        """비용 리포트 PDF 생성"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        
        # 제목
        title = Paragraph("비용 분석 리포트", self.title_style)
        story.append(title)
        story.append(Spacer(1, 0.2 * inch))
        
        # 기간
        period_text = f"기간: {data['period']['start']} ~ {data['period']['end']}"
        period = Paragraph(period_text, self.body_style)
        story.append(period)
        story.append(Spacer(1, 0.3 * inch))
        
        # 총 비용
        total_title = Paragraph("💰 총 비용", self.subtitle_style)
        story.append(total_title)
        
        total_data = [
            ['항목', '금액 (USD)', '비율'],
            ['CPU', f"${data['total']['cpu_cost']:.2f}", f"{data['breakdown']['cpu_percentage']:.1f}%"],
            ['GPU', f"${data['total']['gpu_cost']:.2f}", f"{data['breakdown']['gpu_percentage']:.1f}%"],
            ['메모리', f"${data['total']['memory_cost']:.2f}", f"{data['breakdown']['memory_percentage']:.1f}%"],
            ['총계', f"${data['total']['total_cost']:.2f}", '100%'],
        ]
        
        total_table = Table(total_data, colWidths=[2*inch, 2*inch, 2*inch])
        total_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), self.korean_font),  # 한글 폰트
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#86efac')),
            ('FONTNAME', (0, -1), (-1, -1), self.korean_font),
            ('FONTSIZE', (0, -1), (-1, -1), 14),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(total_table)
        story.append(Spacer(1, 0.3 * inch))
        
        # 요금 정보
        rate_title = Paragraph("📋 요금 정보", self.subtitle_style)
        story.append(rate_title)
        
        rate_data = [
            ['리소스', '단가'],
            ['CPU', f"${data['rates']['cpu_per_hour']:.2f} / hour"],
            ['GPU', f"${data['rates']['gpu_per_hour']:.2f} / hour"],
            ['메모리', f"${data['rates']['memory_per_gb_hour']:.4f} / GB·hour"],
        ]
        
        rate_table = Table(rate_data, colWidths=[3*inch, 3*inch])
        rate_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), self.korean_font),  # 한글 폰트
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(rate_table)
        
        # 푸터
        story.append(Spacer(1, 0.5 * inch))
        footer_text = f"생성 시간: {data['generated_at']}"
        footer = Paragraph(footer_text, ParagraphStyle('Footer', parent=self.body_style, fontName=self.korean_font, fontSize=8, textColor=colors.grey, alignment=TA_RIGHT))
        story.append(footer)
        
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    # ============================================
    # Excel 생성
    # ============================================
    
    def generate_usage_report_excel(self, data):
        """사용량 리포트 Excel 생성"""
        buffer = BytesIO()
        wb = Workbook()
        
        # Sheet 1: 요약
        ws_summary = wb.active
        ws_summary.title = "요약"
        
        # 헤더 스타일
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # 제목
        ws_summary['A1'] = "리소스 사용량 리포트"
        ws_summary['A1'].font = Font(bold=True, size=16, color="1E40AF")
        ws_summary.merge_cells('A1:D1')
        
        # 기간 정보
        ws_summary['A2'] = f"기간: {data['period']['start']} ~ {data['period']['end']}"
        ws_summary.merge_cells('A2:D2')
        
        # 요약 테이블
        ws_summary['A4'] = "항목"
        ws_summary['B4'] = "사용량"
        ws_summary['A4'].fill = header_fill
        ws_summary['A4'].font = header_font
        ws_summary['B4'].fill = header_fill
        ws_summary['B4'].font = header_font
        
        summary_rows = [
            ("CPU 시간 (hours)", data['total']['cpu_hours']),
            ("GPU 시간 (hours)", data['total']['gpu_hours']),
            ("메모리 (GB·hours)", data['total']['memory_gb_hours']),
            ("제출된 작업", data['total']['jobs_submitted']),
            ("완료된 작업", data['total']['jobs_completed']),
            ("실패한 작업", data['total']['jobs_failed']),
        ]
        
        for idx, (label, value) in enumerate(summary_rows, start=5):
            ws_summary[f'A{idx}'] = label
            ws_summary[f'B{idx}'] = value
            if isinstance(value, float):
                ws_summary[f'B{idx}'].number_format = '0.00'
        
        # 비용 테이블
        ws_summary['A12'] = "비용 항목"
        ws_summary['B12'] = "금액 (USD)"
        ws_summary['A12'].fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        ws_summary['A12'].font = header_font
        ws_summary['B12'].fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        ws_summary['B12'].font = header_font
        
        cost_rows = [
            ("CPU 비용", data['costs']['cpu_cost']),
            ("GPU 비용", data['costs']['gpu_cost']),
            ("메모리 비용", data['costs']['memory_cost']),
            ("총 비용", data['costs']['total_cost']),
        ]
        
        for idx, (label, value) in enumerate(cost_rows, start=13):
            ws_summary[f'A{idx}'] = label
            ws_summary[f'B{idx}'] = value
            ws_summary[f'B{idx}'].number_format = '$#,##0.00'
            if label == "총 비용":
                ws_summary[f'A{idx}'].font = Font(bold=True)
                ws_summary[f'B{idx}'].font = Font(bold=True)
        
        # 열 너비 조정
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20
        
        # Sheet 2: 일별 데이터
        ws_daily = wb.create_sheet("일별 데이터")
        
        # 헤더
        headers = ['날짜', 'CPU (hours)', 'GPU (hours)', 'Memory (GB·h)', '제출', '완료', '실패']
        for col, header in enumerate(headers, start=1):
            cell = ws_daily.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # 데이터
        for row_idx, day_data in enumerate(data['daily_data'], start=2):
            ws_daily.cell(row=row_idx, column=1, value=day_data['date'])
            ws_daily.cell(row=row_idx, column=2, value=day_data['cpu_hours'])
            ws_daily.cell(row=row_idx, column=3, value=day_data['gpu_hours'])
            ws_daily.cell(row=row_idx, column=4, value=day_data['memory_gb_hours'])
            ws_daily.cell(row=row_idx, column=5, value=day_data['jobs_submitted'])
            ws_daily.cell(row=row_idx, column=6, value=day_data['jobs_completed'])
            ws_daily.cell(row=row_idx, column=7, value=day_data['jobs_failed'])
            
            # 숫자 포맷
            for col in range(2, 5):
                ws_daily.cell(row=row_idx, column=col).number_format = '0.00'
        
        # 열 너비 조정
        for col in range(1, 8):
            ws_daily.column_dimensions[chr(64 + col)].width = 15
        
        # 차트 추가 (선택적)
        try:
            chart = LineChart()
            chart.title = "일별 리소스 사용량"
            chart.style = 10
            chart.y_axis.title = 'Hours'
            chart.x_axis.title = '날짜'
            
            data_ref = Reference(ws_daily, min_col=2, min_row=1, max_row=len(data['daily_data']) + 1, max_col=3)
            cats = Reference(ws_daily, min_col=1, min_row=2, max_row=len(data['daily_data']) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            
            ws_daily.add_chart(chart, "I2")
        except:
            pass  # 차트 생성 실패 시 무시
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def generate_costs_report_excel(self, data):
        """비용 리포트 Excel 생성"""
        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "비용 분석"
        
        # 헤더 스타일
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # 제목
        ws['A1'] = "비용 분석 리포트"
        ws['A1'].font = Font(bold=True, size=16, color="10B981")
        ws.merge_cells('A1:D1')
        
        # 기간
        ws['A2'] = f"기간: {data['period']['start']} ~ {data['period']['end']}"
        ws.merge_cells('A2:D2')
        
        # 총 비용
        ws['A4'] = "항목"
        ws['B4'] = "금액 (USD)"
        ws['C4'] = "비율 (%)"
        for col in ['A4', 'B4', 'C4']:
            ws[col].fill = header_fill
            ws[col].font = header_font
        
        rows = [
            ("CPU", data['total']['cpu_cost'], data['breakdown']['cpu_percentage']),
            ("GPU", data['total']['gpu_cost'], data['breakdown']['gpu_percentage']),
            ("메모리", data['total']['memory_cost'], data['breakdown']['memory_percentage']),
            ("총계", data['total']['total_cost'], 100),
        ]
        
        for idx, (label, cost, pct) in enumerate(rows, start=5):
            ws[f'A{idx}'] = label
            ws[f'B{idx}'] = cost
            ws[f'C{idx}'] = pct
            ws[f'B{idx}'].number_format = '$#,##0.00'
            ws[f'C{idx}'].number_format = '0.0'
            if label == "총계":
                ws[f'A{idx}'].font = Font(bold=True, size=12)
                ws[f'B{idx}'].font = Font(bold=True, size=12)
                ws[f'C{idx}'].font = Font(bold=True, size=12)
        
        # 요금 정보
        ws['A10'] = "요금 정보"
        ws['A10'].font = Font(bold=True, size=12, color="3B82F6")
        ws.merge_cells('A10:B10')
        
        ws['A11'] = "리소스"
        ws['B11'] = "단가"
        for col in ['A11', 'B11']:
            ws[col].fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            ws[col].font = header_font
        
        ws['A12'] = "CPU"
        ws['B12'] = f"${data['rates']['cpu_per_hour']} / hour"
        ws['A13'] = "GPU"
        ws['B13'] = f"${data['rates']['gpu_per_hour']} / hour"
        ws['A14'] = "메모리"
        ws['B14'] = f"${data['rates']['memory_per_gb_hour']} / GB·hour"
        
        # 열 너비
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        
        # 파이 차트 추가
        try:
            pie = PieChart()
            labels = Reference(ws, min_col=1, min_row=5, max_row=7)
            data_ref = Reference(ws, min_col=2, min_row=4, max_row=7)
            pie.add_data(data_ref, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "비용 분포"
            ws.add_chart(pie, "E2")
        except:
            pass
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer

# 전역 exporter 인스턴스
report_exporter = ReportExporter()
